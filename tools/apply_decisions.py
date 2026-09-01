#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
apply_decisions.py
==================

Read «تصمیم_پانویس_ها.xlsx» back after the author has filled in the
«تصمیم شما» column, and build the final thesis accordingly.

    python3 tools/apply_decisions.py [--out Payannameh_Fatemeh.Bayat-B-v3.docx]

Decisions understood in column «تصمیم شما»:
    اعمال شود    replace the footnote text with «پانویس پیشنهادی»
    بدون تغییر   leave the footnote exactly as it is
    حذف شود      remove the footnote (its marker is removed from the body too)
    (empty)      treated as «بدون تغییر»

Only word/footnotes.xml (and word/document.xml if something is deleted) is
touched, so all the manual formatting in the thesis survives untouched.

Run with --dry-run first to see what would happen.
"""

from __future__ import annotations

import argparse
import copy
import os
import shutil
import zipfile

from lxml import etree
from openpyxl import load_workbook

W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
w = lambda t: f"{{{W}}}{t}"  # noqa: E731

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(REPO, "Payannameh_Fatemeh.Bayat-B-v2.docx")
XLSX = os.path.join(REPO, "تصمیم_پانویس_ها.xlsx")

APPLY, KEEP, DROP = "اعمال شود", "بدون تغییر", "حذف شود"


def read_decisions(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["تصمیم پانویس‌ها"]
    hdr = [c.value for c in ws[1]]
    i_id = hdr.index("ش")
    i_new = hdr.index("پانویس پیشنهادی")
    i_dec = hdr.index("تصمیم شما")
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        fid = row[i_id]
        if fid is None:
            continue
        dec = (row[i_dec] or "").strip()
        out[int(fid)] = (dec, (row[i_new] or "").strip())
    return out


def ensure_ref(note):
    """Give a footnote Word's automatic number if it lacks one."""
    if note.find(".//" + w("footnoteRef")) is not None:
        return False
    para = note.find(w("p"))
    if para is None:
        return False
    tmpl = next((r for r in note.iter(w("r"))
                 if r.find(w("rPr")) is not None), None)
    r = etree.Element(w("r"))
    rPr = etree.SubElement(r, w("rPr"))
    st = etree.SubElement(rPr, w("rStyle"))
    st.set(w("val"), "FootnoteReference")
    if tmpl is not None:
        f = tmpl.find(w("rPr")).find(w("rFonts"))
        if f is not None:
            rPr.append(copy.deepcopy(f))
    etree.SubElement(r, w("footnoteRef"))
    anchor = para.find(w("pPr"))
    (anchor.addnext(r) if anchor is not None else para.insert(0, r))
    return True


def set_text(note, text):
    runs = [r for r in note.iter(w("r"))
            if r.find(w("footnoteRef")) is None and r.find(w("t")) is not None]
    if not runs:
        return
    ts = runs[0].findall(w("t"))
    ts[0].text = " " + text
    ts[0].set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    for extra in ts[1:]:
        extra.text = ""
    for r in runs[1:]:
        for t in r.findall(w("t")):
            t.text = ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.join(
        REPO, "Payannameh_Fatemeh.Bayat-B-v3.docx"))
    ap.add_argument("--xlsx", default=XLSX)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    dec = read_decisions(args.xlsx)
    work = "/tmp/applydec"
    shutil.rmtree(work, ignore_errors=True)
    with zipfile.ZipFile(SRC) as z:
        z.extractall(work)

    fpath = os.path.join(work, "word", "footnotes.xml")
    dpath = os.path.join(work, "word", "document.xml")
    ftree = etree.parse(fpath)
    dtree = etree.parse(dpath)
    froot, droot = ftree.getroot(), dtree.getroot()

    changed, deleted, marks = [], [], []
    for note in froot.findall(w("footnote")):
        if note.get(w("type")):
            continue
        fid = int(note.get(w("id")))
        action, proposed = dec.get(fid, (KEEP, ""))
        cur = "".join(t.text or "" for t in note.iter(w("t"))).strip()

        if action == DROP:
            deleted.append((fid, cur))
            continue
        if ensure_ref(note):
            marks.append(fid)
        if action == APPLY and proposed and proposed != cur:
            set_text(note, proposed)
            changed.append((fid, cur, proposed))

    if args.dry_run:
        print(f"would correct {len(changed)} footnote(s):")
        for fid, o, n in changed:
            print(f"  [{fid:>2}] {o!r} -> {n!r}")
        print(f"would add auto-numbers to: {marks}")
        print(f"would delete: {[f for f, _ in deleted]}")
        return

    # actually delete
    for fid, _ in deleted:
        for note in froot.findall(w("footnote")):
            if note.get(w("id")) == str(fid):
                froot.remove(note)
        for ref in droot.iter(w("footnoteReference")):
            if ref.get(w("id")) == str(fid):
                run = ref.getparent()
                run.getparent().remove(run)

    ftree.write(fpath, xml_declaration=True, encoding="UTF-8",
                standalone=True)
    if deleted:
        dtree.write(dpath, xml_declaration=True, encoding="UTF-8",
                    standalone=True)

    if os.path.exists(args.out):
        os.remove(args.out)
    with zipfile.ZipFile(args.out, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(os.path.join(work, "[Content_Types].xml"),
                "[Content_Types].xml")
        for rd, _, files in os.walk(work):
            for f in files:
                full = os.path.join(rd, f)
                rel = os.path.relpath(full, work).replace(os.sep, "/")
                if rel != "[Content_Types].xml":
                    z.write(full, rel)

    print(f"corrected  : {len(changed)}")
    print(f"auto-number: {marks}")
    print(f"deleted    : {[f for f, _ in deleted]}")
    print("written    :", args.out)


if __name__ == "__main__":
    main()

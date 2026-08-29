# -*- coding: utf-8 -*-
"""MasterThesis v1.1: پانویس از جدول، منابع APA، فهرست، استایل References.

صفحات قبل از فهرست مطالب و صفحهٔ آخر (Payame Noor University …)
از نظر محتوا و شکل دست نمی‌خورند.
"""
import copy
import os
import re
import sys
import zipfile

from lxml import etree
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from apply_v13 import (
    LETTERS, XML_SPACE, bounded_find, insert_run_after_phrase,
    make_fn_ref_run, split_run_at,
)
from apply_v19 import (
    RPR_ORDER, TNR, TITR, LOTUS, ensure, finish_ppr, fonts, make_latin,
    ptext, q, reorder, set_pstyle, setv, style_of,
)

SRC = 'MasterThesis-Fatemeh-Bayat-v1.0.docx'
DST = 'MasterThesis-Fatemeh-Bayat-v1.1.docx'
XLSX = 'Jadval-Panavis-Farsi-Latin.xlsx'
PAGE_CHARS = 2000
W = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'
R = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
REL_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
HYPER_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink'

VARIANTS = {
    'سالکوویس / سالکوسکیس': ['سالکوویس', 'سالکوسکیس', 'سالکوفسکیس'],
    'انجمن روان‌پزشکی آمریکا': ['انجمن روان‌پزشکی آمریکا', 'انجمن روان پزشکی آمریکا'],
    'انجمن روان‌شناسی آمریکا': ['انجمن روان‌شناسی آمریکا', 'انجمن روان شناسی آمریکا'],
    'دی‌سیکو': ['دی‌سیکو', 'دی سیکو'],
    'پیشزچینسکی': ['پیشزچینسکی', 'پیشزینسکی'],
    'تورسون': ['تورسون', 'دورسون'],
    'پاول': ['پاول', 'پاوول'],
    'هایگ': ['هایگ', 'هایگا'],
    'پورهیت': ['پورهیت', 'پوروحیت'],
    'هالدورسون': ['هالدورسون', 'هالدورسن'],
    'چاریکچی‌اوزگول': ['چاریکچی‌اوزگول', 'چاریکچی اوزگول'],
    'موریرا-آلمیدا': ['موریرا-آلمیدا', 'موریرا'],
    'نیکولاس': ['نیکولاس', 'نیکولیچ'],
}

SKIP_FA = {'رو', 'وانگ (Wong)', 'گould'}

DELETE_PREFIXES = (
    'Campisi, J. (2024)',
    'Gladyshev, V. N. (2024)',
    'Fergus, T. A.',
    'Chlan,',
    'Multiple perspectives of spiritual intelligence',
    'Martin, L. L.',
    'King, D. B. (2021)',
    'van Bruggen,',
    'Huang, Y., et al.',
    'Chen, Y., et al.',
    'Kaplan, S., et al.',
    'Gould, R. L., et al.',
    'Soto, C., et al.',
    'Wang, C., et al.',
    'Nikolich',
    'Nordgren, L. F.',
    'World Health Organization. (2020)',
    'World Health Organization. (2024). Ageing and Health.',
    'National Institute on Aging. (2023)',
    'Fulop, T.',
    'Fülöp, T.',
)

REWRITE_BIB = {
    'Havighurst, R. J. (1961)': (
        'Havighurst, R. J. (1961). Successful aging. The Gerontologist, 1(1), 8–13. ',
        'https://doi.org/10.1093/geront/1.1.8',
    ),
    'Suad, Zarina Mat': (
        'Saad, Z. M., Hatta, Z. A., & Mohamad, N. (2010). The impact of spiritual intelligence on the health of the elderly in Malaysia. Asian Social Work and Policy Review, 4(2), 84–97.',
        None,
    ),
    'Moreira-almaida': (
        'Moreira-Almeida, A., & Koenig, H. G. (2006). Retaining the meaning of the words religiousness and spirituality: A commentary on the WHOQOL SRPB group’s cross-cultural study. Social Science & Medicine, 63(4), 843–845.',
        None,
    ),
    'Noyes, R., Jr., Stuart, S., Longley': (
        'Noyes, R., Jr., Stuart, S., Longley, S. L., Langbehn, D. R., & Happel, R. L. (2002). Hypochondriasis and fear of death. Journal of Nervous and Mental Disease, 190(8), 503–509. ',
        'https://doi.org/10.1097/00005053-200208000-00002',
    ),
}

FA_TYPOS = (
    ('اخالقی', 'اخلاقی'),
    ('سالمندانساکن', 'سالمندان ساکن'),
    ('خانهسالمندان', 'خانه سالمندان'),
    ('پژوه انجام', 'پژوهش انجام'),
    ('دستاورهای', 'دستاوردهای'),
)


def footnote_body(latin):
    if 'WHO' in latin:
        return 'WHO'
    if 'APA' in latin:
        return 'APA'
    if '(UN)' in latin:
        return 'UN'
    if '(NIA)' in latin:
        return 'NIA'
    return latin.strip()


def load_pairs():
    wb = load_workbook(XLSX)
    ws = wb.active
    pairs = []
    for typ, fa, la in ws.iter_rows(min_row=2, values_only=True):
        if not fa or not la:
            continue
        fa, la = str(fa).strip(), str(la).strip()
        if fa in SKIP_FA:
            continue
        body = footnote_body(la)
        needles = VARIANTS.get(fa, [fa])
        if ' / ' in fa and fa not in VARIANTS:
            needles = [x.strip() for x in fa.split('/') if x.strip()]
        pairs.append((typ, fa, la, body, needles))
    extra = [
        (['DAS'], 'Death Anxiety Scale (DAS)'),
        (['HAI'], 'Health Anxiety Inventory (HAI)'),
        (['DSM-5'], 'Diagnostic and Statistical Manual of Mental Disorders'),
        (['TMT'], 'Terror Management Theory'),
        (['WHO'], 'WHO'),
        (['APA'], 'APA'),
    ]
    for nd, body in extra:
        pairs.append(('extra', nd[0], body, body, nd))
    pairs.sort(key=lambda x: -max(len(n) for n in x[4]))
    return pairs


def top_paras(body):
    return [el for el in body if el.tag == q('p')]


def find_para(paras, pred):
    for i, p in enumerate(paras):
        if pred(p):
            return i, p
    return None, None


def has_page_break(p):
    ppr = p.find(q('pPr'))
    if ppr is not None and ppr.find(q('pageBreakBefore')) is not None:
        return True
    if p.find('.//' + q('lastRenderedPageBreak')) is not None:
        return True
    for br in p.iter(q('br')):
        if br.get(q('type')) == 'page':
            return True
    return False


def in_table(p):
    parent = p.getparent()
    while parent is not None:
        tag = parent.tag
        if tag == q('tbl') or tag == q('tc'):
            return True
        parent = parent.getparent()
    return False


def skip_for_fn(p):
    st = style_of(p) or ''
    if st.startswith('Heading') or st.startswith('TOC') or st in (
            'Caption', 'TableofFigures', 'Bibliography', 'EnglishText'):
        return True
    t = ptext(p).strip()
    if t.startswith('فهرست') or t.startswith('جدول ') or t.startswith('شکل '):
        return True
    return False


def make_footnote_el(fid, english):
    el = etree.Element(q('footnote'))
    el.set(q('id'), str(fid))
    p = etree.SubElement(el, q('p'))
    ppr = etree.SubElement(p, q('pPr'))
    ps = etree.SubElement(ppr, q('pStyle'))
    ps.set(q('val'), 'FootnoteText')
    setv(ppr, 'bidi', val='0')
    setv(ppr, 'jc', val='both')
    rpr = etree.SubElement(ppr, q('rPr'))
    fonts(rpr, TNR, TNR)
    setv(rpr, 'rtl', val='0')
    setv(rpr, 'cs', val='0')
    setv(rpr, 'sz', val='18')
    setv(rpr, 'szCs', val='18')
    lang = ensure(rpr, 'lang')
    lang.set(q('val'), 'en-US')
    finish_ppr(ppr)
    rmark = etree.SubElement(p, q('r'))
    rp = etree.SubElement(rmark, q('rPr'))
    st = etree.SubElement(rp, q('rStyle'))
    st.set(q('val'), 'FootnoteReference')
    fonts(rp, TNR, TNR)
    etree.SubElement(rmark, q('footnoteRef'))
    tr = etree.SubElement(p, q('r'))
    rpr = etree.SubElement(tr, q('rPr'))
    fonts(rpr, TNR, TNR)
    setv(rpr, 'sz', val='18')
    setv(rpr, 'szCs', val='18')
    setv(rpr, 'rtl', val='0')
    setv(rpr, 'cs', val='0')
    lang = ensure(rpr, 'lang')
    lang.set(q('val'), 'en-US')
    te = etree.SubElement(tr, q('t'))
    te.set(XML_SPACE, 'preserve')
    te.text = ' ' + english
    return el


def latin_fn_style(fn_root):
    for f in fn_root.findall(q('footnote')):
        if f.get(q('type')):
            continue
        for p in f.findall(q('p')):
            ppr = p.find(q('pPr'))
            if ppr is None:
                ppr = etree.Element(q('pPr'))
                p.insert(0, ppr)
            setv(ppr, 'bidi', val='0')
            setv(ppr, 'jc', val='both')
            rpr = ppr.find(q('rPr'))
            if rpr is None:
                rpr = etree.SubElement(ppr, q('rPr'))
            fonts(rpr, TNR, TNR)
            setv(rpr, 'rtl', val='0')
            setv(rpr, 'cs', val='0')
            setv(rpr, 'sz', val='18')
            setv(rpr, 'szCs', val='18')
            lang = ensure(rpr, 'lang')
            lang.set(q('val'), 'en-US')
            finish_ppr(ppr)


def collect_matches(plain, pairs):
    found = []
    occupied = [False] * (len(plain) + 1)
    for typ, fa, la, body, needles in pairs:
        best = None
        for n in needles:
            hits = bounded_find(plain, n)
            if not hits and (' ' in n or '\u200c' in n) and n in plain:
                i = plain.find(n)
                class M:
                    pass
                m = M()
                m.start = lambda i=i: i
                m.end = lambda i=i, n=n: i + len(n)
                m.group = lambda n=n: n
                hits = [m]
            if hits:
                h = hits[0]
                if best is None or h.start() < best[0]:
                    best = (h.start(), h.end(), n, body)
        if best is None:
            continue
        s, e, n, body = best
        if any(occupied[s:e]):
            continue
        for i in range(s, e):
            occupied[i] = True
        found.append((s, e, n, body))
    # وانگ: ۲۰۰۸ = Wong وگرنه Wang
    for m in bounded_find(plain, 'وانگ'):
        s, e = m.start(), m.end()
        if any(occupied[s:e]):
            continue
        nxt = plain[e:e + 12]
        body = 'Wong' if re.search(r'،\s*۲۰۰۸', nxt) or re.search(r'،\s*2008', nxt) else 'Wang'
        for i in range(s, e):
            occupied[i] = True
        found.append((s, e, 'وانگ', body))
    # رو و کان → Rowe سپس Kahn
    i = plain.find('رو و کان')
    if i >= 0:
        found.append((i, i + 2, 'رو', 'Rowe'))
        j = i + len('رو و ')
        found.append((j, j + 3, 'کان', 'Kahn'))
    found.sort(key=lambda x: x[0])
    # unique by body, first occurrence
    seen = set()
    uniq = []
    for item in found:
        if item[3] in seen:
            continue
        seen.add(item[3])
        uniq.append(item)
    return uniq


def add_footnotes(body, fn_root, pairs, i_chek, i_fa):
    kids = list(body)
    used = set()
    chars = 0
    added = 0
    nxt = 1
    for idx in range(i_chek, i_fa):
        el = kids[idx]
        if el.tag != q('p'):
            continue
        p = el
        if has_page_break(p):
            used = set()
            chars = 0
        if chars >= PAGE_CHARS:
            used = set()
            chars = 0
        txt = ptext(p)
        chars += len(txt)
        if skip_for_fn(p) or in_table(p):
            continue
        matches = collect_matches(txt, pairs)
        todo = [(n, body_en) for s, e, n, body_en in matches if body_en not in used]
        todo.reverse()
        for needle, body_en in todo:
            if body_en in used:
                continue
            fid = nxt
            nxt += 1
            fn_root.append(make_footnote_el(fid, body_en))
            run = make_fn_ref_run(fid)
            if insert_run_after_phrase(p, needle, run):
                used.add(body_en)
                added += 1
            else:
                fn_root.remove(fn_root.findall(q('footnote'))[-1])
                nxt -= 1
    return added


def add_style(styles, sid, name, latin=True, hanging=True):
    existing = None
    for s in styles.findall(q('style')):
        if s.get(q('styleId')) == sid:
            existing = s
            break
    if existing is None:
        s = etree.SubElement(styles, q('style'))
        s.set(q('type'), 'paragraph')
        s.set(q('customStyle'), '1')
        s.set(q('styleId'), sid)
        nm = etree.SubElement(s, q('name'))
        nm.set(q('val'), name)
        based = etree.SubElement(s, q('basedOn'))
        based.set(q('val'), 'Normal')
        etree.SubElement(s, q('qFormat'))
    else:
        s = existing
    ppr = s.find(q('pPr'))
    if ppr is None:
        ppr = etree.SubElement(s, q('pPr'))
    rpr = s.find(q('rPr'))
    if rpr is None:
        rpr = etree.SubElement(s, q('rPr'))
    if latin:
        make_latin(ppr, rpr)
        fonts(rpr, TNR, TNR)
        setv(rpr, 'sz', val='24')
        setv(rpr, 'szCs', val='24')
        if hanging:
            ind = ensure(ppr, 'ind')
            ind.set(q('left'), '720')
            ind.set(q('hanging'), '720')
            if q('right') in ind.attrib:
                del ind.attrib[q('right')]
    else:
        setv(ppr, 'bidi', val='1')
        setv(ppr, 'jc', val='right')
        fonts(rpr, TNR, LOTUS)
        setv(rpr, 'sz', val='24')
        setv(rpr, 'szCs', val='24')
        rtl = ensure(rpr, 'rtl')
        if q('val') in rtl.attrib:
            del rtl.attrib[q('val')]
        cs = ensure(rpr, 'cs')
        if q('val') in cs.attrib:
            del cs.attrib[q('val')]
        if hanging:
            ind = ensure(ppr, 'ind')
            ind.set(q('right'), '720')
            ind.set(q('hanging'), '720')
            if q('left') in ind.attrib:
                del ind.attrib[q('left')]
    sp = ensure(ppr, 'spacing')
    sp.set(q('after'), '80')
    sp.set(q('line'), '276')
    sp.set(q('lineRule'), 'auto')
    finish_ppr(ppr)
    reorder(rpr, RPR_ORDER)
    return s


def add_rel(rels, rid, url):
    rel = etree.SubElement(rels, '{%s}Relationship' % REL_NS)
    rel.set('Id', rid)
    rel.set('Type', HYPER_TYPE)
    rel.set('Target', url)
    rel.set('TargetMode', 'External')


def next_rid(rels):
    n = 0
    for r in rels:
        i = r.get('Id') or ''
        if i.startswith('rId'):
            try:
                n = max(n, int(i[3:]))
            except ValueError:
                pass
    return n + 1


def clean_bib_text(t):
    t = t.replace('\\t "_new"', '').replace('\\t "_blank"', '')
    t = t.replace('\t "_new"', '').replace('\t "_blank"', '')
    t = re.sub(r'World Health Organization – [^\n]*', '', t)
    t = re.sub(r'UNFPA Iran – [^\n]*', '', t)
    t = re.sub(r'(https://doi\.org/[^\s\\]+)\1+', r'\1', t)
    t = re.sub(r'(https://www\.[^\s\\]+)\1+', r'\1', t)
    t = re.sub(r'(https://iran\.unfpa\.org/[^\s\\]+)\1+', r'\1', t)
    t = re.sub(r'\\t\s*"_[a-z]+"', '', t)
    t = re.sub(r'(https://doi\.org/[^\s]+)\s*https://doi\.org/[^\s]+', r'\1', t)
    t = re.sub(r'(?<![\s])(https://)', r' \1', t)
    t = t.replace('(14 thed)', '(14th ed.)').replace('(14 thed).', '(14th ed.).')
    t = re.sub(r'\s+', ' ', t).strip()
    t = re.sub(r'\s+\.', '.', t)
    return t


def extract_url(t):
    m = re.search(r'(https://doi\.org/[^\s]+)', t)
    if m:
        return m.group(1).rstrip('.,;')
    m = re.search(r'(https://[^\s]+)', t)
    if m:
        return m.group(1).rstrip('.,;')
    return None


def set_para_text_latin(p, text, rid=None, url=None):
    ppr = p.find(q('pPr'))
    for child in list(p):
        if child is not ppr:
            p.remove(child)
    if ppr is None:
        ppr = etree.Element(q('pPr'))
        p.insert(0, ppr)
    set_pstyle(p, 'References')
    make_latin(ppr, None)
    finish_ppr(ppr)
    display = text
    if url and url in display:
        display = display.replace(url, '').rstrip() + ' '
    r = etree.SubElement(p, q('r'))
    rpr = etree.SubElement(r, q('rPr'))
    fonts(rpr, TNR, TNR)
    setv(rpr, 'rtl', val='0')
    setv(rpr, 'cs', val='0')
    te = etree.SubElement(r, q('t'))
    te.set(XML_SPACE, 'preserve')
    te.text = display if rid and url else text
    if rid and url:
        hl = etree.SubElement(p, q('hyperlink'))
        hl.set(R + 'id', rid)
        hl.set(q('history'), '1')
        hr = etree.SubElement(hl, q('r'))
        hpr = etree.SubElement(hr, q('rPr'))
        st = etree.SubElement(hpr, q('rStyle'))
        st.set(q('val'), 'Hyperlink')
        fonts(hpr, TNR, TNR)
        setv(hpr, 'rtl', val='0')
        ht = etree.SubElement(hr, q('t'))
        ht.text = url


def rewrite_sources(body, rels, i_fa, i_lat, i_abs):
    kids = list(body)
    deleted = 0
    fixed = 0
    rid_n = next_rid(rels)
    url_rids = {}
    for rel in rels:
        if rel.get('Type') == HYPER_TYPE:
            url_rids[rel.get('Target')] = rel.get('Id')

    def rid_for(url):
        nonlocal rid_n
        if not url:
            return None
        if url in url_rids:
            return url_rids[url]
        rid = 'rId%d' % rid_n
        rid_n += 1
        add_rel(rels, rid, url)
        url_rids[url] = rid
        return rid

    # Persian
    for el in kids[i_fa + 1:i_lat]:
        if el.tag != q('p'):
            continue
        t = ptext(el)
        if not t.strip():
            continue
        nt = t
        for a, b in FA_TYPOS:
            nt = nt.replace(a, b)
        ts = list(el.iter(q('t')))
        if ts and nt != t:
            ts[0].text = nt
            for x in ts[1:]:
                x.text = ''
            fixed += 1
        set_pstyle(el, 'PersianReferences')
        ppr = el.find(q('pPr'))
        if ppr is None:
            ppr = etree.Element(q('pPr'))
            el.insert(0, ppr)
        setv(ppr, 'bidi', val='1')
        setv(ppr, 'jc', val='both')
        finish_ppr(ppr)

    # Latin
    to_del = []
    for el in kids[i_lat + 1:i_abs]:
        if el.tag != q('p'):
            continue
        t = ptext(el).strip()
        if not t:
            continue
        if any(t.startswith(px) for px in DELETE_PREFIXES):
            # Nordgren concat → keep Noyes via REWRITE on same para
            if t.startswith('Nordgren') and 'Noyes,' in t:
                key = 'Noyes, R., Jr., Stuart, S., Longley'
                text, url = REWRITE_BIB[key]
                rid = rid_for(url)
                set_para_text_latin(el, text, rid, url)
                fixed += 1
                continue
            to_del.append(el)
            deleted += 1
            continue
        rew = None
        for key, val in REWRITE_BIB.items():
            if key in t[:80] or t.startswith(key[:20]):
                rew = val
                break
        if rew:
            text, url = rew
            rid = rid_for(url)
            set_para_text_latin(el, text, rid, url)
            fixed += 1
            continue
        nt = clean_bib_text(t)
        url = extract_url(nt)
        rid = rid_for(url) if url else None
        set_para_text_latin(el, nt, rid, url)
        fixed += 1
    for el in to_del:
        parent = el.getparent()
        if parent is not None:
            parent.remove(el)
    return deleted, fixed


def add_seq_runs(p, prefix, seq_name, num, rest):
    """Replace caption runs with PREFIX + SEQ + num + rest."""
    ppr = p.find(q('pPr'))
    keep = []
    if ppr is not None:
        keep.append(ppr)
    for b in p.findall(q('bookmarkStart')):
        keep.append(b)
    for child in list(p):
        if child not in keep and child.tag not in (q('bookmarkStart'), q('pPr')):
            if child.tag == q('bookmarkEnd'):
                keep.append(child)
                continue
            p.remove(child)
    def run_t(text):
        r = etree.SubElement(p, q('r'))
        te = etree.SubElement(r, q('t'))
        te.text = text
        return r
    run_t(prefix)
    r1 = etree.SubElement(p, q('r'))
    etree.SubElement(r1, q('fldChar')).set(q('fldCharType'), 'begin')
    r2 = etree.SubElement(p, q('r'))
    it = etree.SubElement(r2, q('instrText'))
    it.set(XML_SPACE, 'preserve')
    it.text = ' SEQ %s \\* ARABIC ' % seq_name
    r3 = etree.SubElement(p, q('r'))
    etree.SubElement(r3, q('fldChar')).set(q('fldCharType'), 'separate')
    run_t(num)
    r5 = etree.SubElement(p, q('r'))
    etree.SubElement(r5, q('fldChar')).set(q('fldCharType'), 'end')
    run_t(rest)
    for b in p.findall(q('bookmarkEnd')):
        p.append(b)


def fix_captions(body):
    n = 0
    for p in body.iter(q('p')):
        if style_of(p) != 'Caption':
            continue
        t = ptext(p).strip()
        if t.startswith('جدول ۴-۱-') and 'جنسیت' in t and not any(
                'SEQ' in (it.text or '') for it in p.iter(q('instrText'))):
            add_seq_runs(p, 'جدول ۴-', 'جدول', '1', '- شاخص‌های توصیفی مربوط به جنسیت')
            n += 1
        m = re.match(r'^شکل\s+([۰-۹0-9]+)[-–]\s*(.*)$', t)
        if m and not any('SEQ' in (it.text or '') for it in p.iter(q('instrText'))):
            add_seq_runs(p, 'شکل ', 'شکل', m.group(1), '- ' + m.group(2))
            n += 1
    return n


def toc_field_para(instr, style):
    p = etree.Element(q('p'))
    ppr = etree.SubElement(p, q('pPr'))
    ps = etree.SubElement(ppr, q('pStyle'))
    ps.set(q('val'), style)
    r1 = etree.SubElement(p, q('r'))
    fc = etree.SubElement(r1, q('fldChar'))
    fc.set(q('fldCharType'), 'begin')
    fc.set(q('dirty'), 'true')
    r2 = etree.SubElement(p, q('r'))
    it = etree.SubElement(r2, q('instrText'))
    it.set(XML_SPACE, 'preserve')
    it.text = instr
    r3 = etree.SubElement(p, q('r'))
    etree.SubElement(r3, q('fldChar')).set(q('fldCharType'), 'separate')
    r4 = etree.SubElement(p, q('r'))
    t = etree.SubElement(r4, q('t'))
    t.text = ' '
    r5 = etree.SubElement(p, q('r'))
    etree.SubElement(r5, q('fldChar')).set(q('fldCharType'), 'end')
    return p


def fix_toc(body):
    n = 0
    for p in body.iter(q('p')):
        for it in p.iter(q('instrText')):
            if it.text and 'TOC \\o "1-4"' in it.text:
                it.text = it.text.replace('TOC \\o "1-4"', 'TOC \\o "1-5"')
                n += 1
        for fc in p.iter(q('fldChar')):
            if fc.get(q('fldCharType')) == 'begin':
                fc.set(q('dirty'), 'true')
    # فهرست اشکال after فهرست جداول block
    kids = list(body)
    i_fig = None
    i_tbl_head = None
    i_sect = None
    for i, el in enumerate(kids):
        if el.tag != q('p'):
            continue
        t = ptext(el).strip()
        if t == 'فهرست جداول':
            i_tbl_head = i
        if t == 'فهرست اشکال':
            i_fig = i
        ppr = el.find(q('pPr'))
        if ppr is not None and ppr.find(q('sectPr')) is not None and i_tbl_head is not None and i_sect is None and i > i_tbl_head:
            i_sect = i
    if i_fig is None and i_sect is not None:
        h = etree.Element(q('p'))
        ppr = etree.SubElement(h, q('pPr'))
        etree.SubElement(ppr, q('bidi'))
        jc = etree.SubElement(ppr, q('jc'))
        jc.set(q('val'), 'center')
        r = etree.SubElement(h, q('r'))
        rpr = etree.SubElement(r, q('rPr'))
        fonts(rpr, TNR, TITR)
        etree.SubElement(rpr, q('b'))
        etree.SubElement(rpr, q('bCs'))
        te = etree.SubElement(r, q('t'))
        te.text = 'فهرست اشکال'
        fld = toc_field_para(' TOC \\h \\z \\c "شکل" ', 'TableofFigures')
        kids[i_sect].addprevious(h)
        kids[i_sect].addprevious(fld)
        n += 1
    return n


def freeze_hash(el):
    return etree.tostring(el, encoding='unicode')


def build():
    pairs = load_pairs()
    print('pairs', len(pairs))
    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    zin.close()
    doc = etree.fromstring(parts['word/document.xml'])
    fn_root = etree.fromstring(parts['word/footnotes.xml'])
    styles = etree.fromstring(parts['word/styles.xml'])
    rels = etree.fromstring(parts['word/_rels/document.xml.rels'])
    settings = etree.fromstring(parts['word/settings.xml'])
    body = doc[0]
    kids = list(body)

    def idx_of(text, style=None):
        for i, el in enumerate(kids):
            if el.tag != q('p'):
                continue
            t = ptext(el).strip()
            if t == text or t.rstrip() == text:
                if style is None or style_of(el) == style:
                    return i
        return None

    i_toc = idx_of('فهرست مطالب')
    i_chek = None
    for i, el in enumerate(kids):
        if el.tag == q('p') and (style_of(el) or '') == 'Heading1' and ptext(el).strip().startswith('چکیده'):
            i_chek = i
            break
    i_fa = idx_of('منابع فارسی', 'Heading1')
    i_lat = idx_of('منابع لاتین', 'Heading1')
    i_abs = idx_of('ABSTRACT', 'Heading1')
    i_pnu = None
    for i, el in enumerate(kids):
        if el.tag == q('p') and 'Payame Noor University' in ptext(el):
            i_pnu = i
            break
    print('idx', dict(toc=i_toc, chek=i_chek, fa=i_fa, lat=i_lat, abs=i_abs, pnu=i_pnu))
    front_xml = [freeze_hash(el) for el in kids[:i_toc]]
    last_xml = [freeze_hash(el) for el in kids[i_pnu:]]

    add_style(styles, 'References', 'References', latin=True)
    add_style(styles, 'PersianReferences', 'Persian References', latin=False)

    nfn = add_footnotes(body, fn_root, pairs, i_chek, i_fa)
    print('footnotes added', nfn)
    latin_fn_style(fn_root)

    kids = list(body)
    i_fa = idx_of('منابع فارسی', 'Heading1')
    i_lat = idx_of('منابع لاتین', 'Heading1')
    i_abs = idx_of('ABSTRACT', 'Heading1')
    deleted, fixed = rewrite_sources(body, rels, i_fa, i_lat, i_abs)
    print('bib deleted', deleted, 'fixed', fixed)

    ncap = fix_captions(body)
    print('captions', ncap)
    ntoc = fix_toc(body)
    print('toc', ntoc)

    if settings.find(q('updateFields')) is None:
        uf = etree.SubElement(settings, q('updateFields'))
        uf.set(q('val'), 'true')
    else:
        settings.find(q('updateFields')).set(q('val'), 'true')

    # freeze check
    kids2 = list(body)
    i_toc2 = None
    i_pnu2 = None
    for i, el in enumerate(kids2):
        if el.tag == q('p') and ptext(el).strip() == 'فهرست مطالب':
            i_toc2 = i
        if el.tag == q('p') and 'Payame Noor University' in ptext(el):
            i_pnu2 = i
    assert i_toc2 is not None and i_pnu2 is not None
    front2 = [freeze_hash(el) for el in kids2[:i_toc2]]
    last2 = [freeze_hash(el) for el in kids2[i_pnu2:]]
    if front2 != front_xml:
        print('ERROR front matter changed', len(front_xml), len(front2))
        for a, b in zip(front_xml, front2):
            if a != b:
                print('DIFF', a[:80], '||', b[:80])
                break
        raise SystemExit(2)
    if last2 != last_xml:
        print('ERROR last page changed', len(last_xml), len(last2))
        raise SystemExit(3)
    print('freeze OK front', len(front2), 'last', len(last2))

    used = {fr.get(q('id')) for fr in doc.iter(q('footnoteReference'))}
    defined = {f.get(q('id')) for f in fn_root.findall(q('footnote')) if not f.get(q('type'))}
    print('fn defined', len(defined), 'used', len(used), 'equal', defined == used)

    parts['word/document.xml'] = etree.tostring(
        doc, xml_declaration=True, encoding='UTF-8', standalone=True)
    parts['word/footnotes.xml'] = etree.tostring(
        fn_root, xml_declaration=True, encoding='UTF-8', standalone=True)
    parts['word/styles.xml'] = etree.tostring(
        styles, xml_declaration=True, encoding='UTF-8', standalone=True)
    parts['word/_rels/document.xml.rels'] = etree.tostring(
        rels, xml_declaration=True, encoding='UTF-8', standalone=True)
    parts['word/settings.xml'] = etree.tostring(
        settings, xml_declaration=True, encoding='UTF-8', standalone=True)
    with zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED) as zout:
        for k, v in parts.items():
            zout.writestr(k, v)
    print('نوشته شد:', DST, os.path.getsize(DST))


if __name__ == '__main__':
    build()

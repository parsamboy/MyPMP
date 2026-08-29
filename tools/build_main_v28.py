# -*- coding: utf-8 -*-
"""v2.8: حذف پانویس از متن + جدول یکتای فارسی↔لاتین برای بازیابی."""
import json
import os
import sys
import zipfile

from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from apply_v19 import q

SRC = 'Payannameh-Fatemeh-Bayat-v2.7.docx'
DST = 'Payannameh-Fatemeh-Bayat-v2.8.docx'
XLSX = 'Jadval-Panavis-Farsi-Latin.xlsx'
CSV = 'Jadval-Panavis-Farsi-Latin.csv'
JSONP = 'Jadval-Panavis-Farsi-Latin.json'

# فارسی در متن پایان‌نامه → لاتین اصلی (بدون تکرار)
# ترتیب: الفبای لاتین
ROWS = [
    ('شخص', 'آبراموویتز', 'Abramowitz'),
    ('شخص', 'آچلی', 'Atchley'),
    ('شخص', 'آدلر', 'Adler'),
    ('شخص', 'آسموندسون', 'Asmundson'),
    ('شخص', 'آیزاکوویتز', 'Isaacowitz'),
    ('سازمان', 'انجمن روان‌پزشکی آمریکا', 'American Psychiatric Association (APA)'),
    ('سازمان', 'انجمن روان‌شناسی آمریکا', 'American Psychological Association (APA)'),
    ('شخص', 'اتچیسون', 'Atchison'),
    ('شخص', 'اریکسون', 'Erikson'),
    ('شخص', 'استاوروا', 'Stavrova'),
    ('شخص', 'استرنبرگ', 'Sternberg'),
    ('شخص', 'استوارت', 'Stuart'),
    ('شخص', 'انگل', 'Engel'),
    ('شخص', 'ایوراچ', 'Iverach'),
    ('شخص', 'ایشیک', 'Işık'),
    ('مفهوم', 'اینفلامجینگ', 'Inflammaging'),
    ('مفهوم', 'اثر مثبت‌نگری', 'positivity effect'),
    ('شخص', 'برادوک', 'Braddock'),
    ('شخص', 'بک', 'Beck'),
    ('شخص', 'بالبی', 'Bowlby'),
    ('شخص', 'بولمایر', 'Bohlmeijer'),
    ('مقیاس', 'پرسشنامه اضطراب سلامتی', 'Health Anxiety Inventory (HAI)'),
    ('مقیاس', 'پرسشنامه اضطراب مرگ', 'Death Anxiety Scale (DAS)'),
    ('شخص', 'پاپالیا', 'Papalia'),
    ('شخص', 'پاول', 'Powell'),
    ('شخص', 'پورهیت', 'Purohit'),
    ('شخص', 'پینتو', 'Pinto'),
    ('شخص', 'پیشزچینسکی', 'Pyszczynski'),
    ('شخص', 'تامر', 'Tomer'),
    ('شخص', 'توبین', 'Tobin'),
    ('شخص', 'تورسون', 'Thorson'),
    ('شخص', 'تمپلر', 'Templer'),
    ('شخص', 'تیلور', 'Taylor'),
    ('شخص', 'جیمز', 'William James'),
    ('شخص', 'جین', 'Jain'),
    ('مفهوم', 'ژرونتولوژی', 'Gerontology'),
    ('شخص', 'داود', 'Dowd'),
    ('شخص', 'داینر', 'Diener'),
    ('شخص', 'دسی', 'Deci'),
    ('شخص', 'دی‌سیکو', 'DeCicco'),
    ('شخص', 'رایان', 'Ryan'),
    ('شخص', 'رید', 'Reed'),
    ('شخص', 'ریف', 'Ryff'),
    ('شخص', 'زنگ', 'Zeng'),
    ('شخص', 'زوهار', 'Zohar'),
    ('شخص', 'سالکوویس / سالکوسکیس', 'Salkovskis'),
    ('شخص', 'سالووی', 'Salovey'),
    ('سازمان', 'سازمان جهانی بهداشت', 'World Health Organization (WHO)'),
    ('سازمان', 'سازمان ملل متحد', 'United Nations (UN)'),
    ('شخص', 'سوتو', 'Soto'),
    ('شخص', 'سعاد', 'Mat Saad'),
    ('شخص', 'شی', 'Schaie'),
    ('شخص', 'فرانچسکی', 'Franceschi'),
    ('شخص', 'فرانکل', 'Frankl'),
    ('شخص', 'فرگوس', 'Fergus'),
    ('شخص', 'فرنس', 'Franz'),
    ('شخص', 'فروید', 'Freud'),
    ('شخص', 'فلوریان', 'Florian'),
    ('شخص', 'فولپ', 'Fülöp'),
    ('شخص', 'کارستنسن', 'Carstensen'),
    ('شخص', 'کاپلان', 'Kaplan'),
    ('شخص', 'کامپیزی', 'Campisi'),
    ('شخص', 'کامینگ', 'Cumming'),
    ('شخص', 'کان', 'Kahn'),
    ('شخص', 'کرجسی', 'Krejcie'),
    ('شخص', 'کینگ', 'King'),
    ('شخص', 'کونیگ', 'Koenig'),
    ('شخص', 'کیز', 'Keyes'),
    ('شخص', 'کیکاس', 'Kikas'),
    ('شخص', 'کیونیو', 'Kivnick'),
    ('شخص', 'گلادیشف', 'Gladyshev'),
    ('شخص', 'گould', 'Gould'),
    ('شخص', 'گرینبرگ', 'Greenberg'),
    ('شخص', 'گولد', 'Gould'),
    ('شخص', 'لوپز-اوتین', 'López-Otín'),
    ('شخص', 'لوین', 'Levin'),
    ('شخص', 'لوینسون', 'Levinson'),
    ('شخص', 'لیو', 'Liu'),
    ('شخص', 'مارشال', 'Marshall'),
    ('شخص', 'مارتین', 'Martin'),
    ('شخص', 'مارتورل', 'Martorell'),
    ('شخص', 'مازلو', 'Maslow'),
    ('شخص', 'مایر', 'Mayer'),
    ('شخص', 'مایکلز', 'Mikels'),
    ('شخص', 'مک‌آدامز', 'McAdams'),
    ('شخص', 'منزیس', 'Menzies'),
    ('شخص', 'مورگان', 'Morgan'),
    ('شخص', 'موریرا-آلمیدا', 'Moreira-Almeida'),
    ('شخص', 'میکولینسر', 'Mikulincer'),
    ('سازمان', 'مؤسسه ملی سالمندی', 'National Institute on Aging (NIA)'),
    ('نظریه', 'نظریه ایمنی', 'Immunological Theory'),
    ('نظریه', 'نظریه انتخاب اجتماعی-هیجانی', 'Socioemotional Selectivity Theory'),
    ('نظریه', 'نظریه تداوم', 'Continuity Theory'),
    ('نظریه', 'نظریه فعالیت', 'Activity Theory'),
    ('نظریه', 'نظریه مدیریت وحشت', 'Terror Management Theory'),
    ('نظریه', 'نظریه مدیریت معنا', 'Meaning Management Theory'),
    ('نظریه', 'نظریه عدم تعهد', 'Disengagement Theory'),
    ('شخص', 'نیمیر', 'Neimeyer'),
    ('شخص', 'نیوگارتن', 'Neugarten'),
    ('شخص', 'نیکولاس', 'Nikolich-Žugich'),
    ('شخص', 'نوردگرن', 'Nordgren'),
    ('شخص', 'نویز', 'Noyes'),
    ('شخص', 'ون بروگن', 'van Bruggen'),
    ('شخص', 'وسترهاف', 'Westerhof'),
    ('شخص', 'وارویک', 'Warwick'),
    ('شخص', 'والنتاینر', 'Valentiner'),
    ('شخص', 'وانگ', 'Wang'),
    ('شخص', 'وانگ (Wong)', 'Wong'),
    ('شخص', 'هالدورسون', 'Halldórsson'),
    ('شخص', 'هاردینگ', 'Harding'),
    ('شخص', 'هاویگهرست', 'Havighurst'),
    ('شخص', 'هایگ', 'Haigh'),
    ('شخص', 'هوانگ', 'Huang'),
    ('شخص', 'هنری', 'Henry'),
    ('شخص', 'یالوم', 'Yalom'),
    ('شخص', 'یانکر', 'Yonker'),
    ('شخص', 'یونگ', 'Jung'),
    ('شخص', 'چالن', 'Chlan'),
    ('شخص', 'چان', 'Chan'),
    ('شخص', 'چارلز', 'Charles'),
    ('شخص', 'چاریکچی‌اوزگول', 'Çarıkçı-Özgül'),
    ('شخص', 'چن', 'Chen'),
    ('شخص', 'ژائو', 'Zhao'),
    ('شخص', 'الیاسون', 'Eliason'),
    ('شخص', 'رو', 'Rowe'),
    ('مقیاس', 'DSM-5', 'Diagnostic and Statistical Manual of Mental Disorders'),
]


def strip_footnotes(doc, fn_root):
    n = 0
    for el in list(doc.iter(q('footnoteReference'))):
        r = el.getparent()
        if r is None:
            continue
        parent = r.getparent()
        if parent is None:
            continue
        # if run has only the footnote mark, drop the run
        has_text = any((t.text or '').strip() for t in r.iter(q('t')))
        if has_text:
            r.remove(el)
        else:
            parent.remove(r)
        n += 1
    # keep only separator footnotes
    for f in list(fn_root.findall(q('footnote'))):
        if not f.get(q('type')):
            fn_root.remove(f)
    return n


def write_table(rows):
    # unique by (fa, la)
    seen = set()
    clean = []
    for typ, fa, la in rows:
        if fa == 'گould':
            continue
        key = (fa, la)
        if key in seen:
            continue
        seen.add(key)
        clean.append((typ, fa, la))
    # csv
    with open(CSV, 'w', encoding='utf-8-sig') as f:
        f.write('نوع,فارسی در متن,لاتین / انگلیسی\n')
        for typ, fa, la in clean:
            f.write('%s,%s,%s\n' % (typ, fa.replace(',', '،'), la.replace(',', '،')))
    # json
    with open(JSONP, 'w', encoding='utf-8') as f:
        json.dump(
            [{'type': t, 'fa': fa, 'la': la} for t, fa, la in clean],
            f, ensure_ascii=False, indent=2)
    # xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = 'Farsi-Latin'
    ws.sheet_view.rightToLeft = True
    headers = ['نوع', 'فارسی در متن پایان‌نامه', 'لاتین / انگلیسی']
    head_fill = PatternFill('solid', fgColor='1B4F72')
    head_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'))
    for i, h in enumerate(headers, 1):
        c = ws.cell(1, i, h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin
    fills = {
        'شخص': PatternFill('solid', fgColor='EAF2F8'),
        'سازمان': PatternFill('solid', fgColor='FEF9E7'),
        'نظریه': PatternFill('solid', fgColor='E8F8F5'),
        'مفهوم': PatternFill('solid', fgColor='F5EEF8'),
        'مقیاس': PatternFill('solid', fgColor='FDEDEC'),
    }
    for r, (typ, fa, la) in enumerate(clean, 2):
        vals = (typ, fa, la)
        fill = fills.get(typ)
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v)
            c.font = Font(name='Calibri', size=12)
            c.alignment = Alignment(horizontal='right' if i < 3 else 'left',
                                    vertical='center', wrap_text=True)
            c.border = thin
            if fill:
                c.fill = fill
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 48
    ws.auto_filter.ref = 'A1:C%d' % (len(clean) + 1)
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 22
    wb.save(XLSX)
    return clean


def build():
    zin = zipfile.ZipFile(SRC)
    parts = {n: zin.read(n) for n in zin.namelist()}
    zin.close()
    doc = etree.fromstring(parts['word/document.xml'])
    fn_root = etree.fromstring(parts['word/footnotes.xml'])
    n = strip_footnotes(doc, fn_root)
    print('removed refs', n)
    left = list(doc.iter(q('footnoteReference')))
    print('left refs', len(left))
    nfn = len([f for f in fn_root.findall(q('footnote')) if not f.get(q('type'))])
    print('custom fn left', nfn)

    parts['word/document.xml'] = etree.tostring(
        doc, xml_declaration=True, encoding='UTF-8', standalone=True)
    parts['word/footnotes.xml'] = etree.tostring(
        fn_root, xml_declaration=True, encoding='UTF-8', standalone=True)
    with zipfile.ZipFile(DST, 'w', zipfile.ZIP_DEFLATED) as zout:
        for k, v in parts.items():
            zout.writestr(k, v)
    clean = write_table(ROWS)
    print('table rows', len(clean))
    print('نوشته شد:', DST, os.path.getsize(DST))
    print('جدول:', XLSX, CSV)


if __name__ == '__main__':
    build()

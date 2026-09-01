#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_decision_xlsx.py
======================

Produce «تصمیم_پانویس_ها.xlsx» — a compact, print-ready decision sheet for all
85 footnotes of Payannameh_Fatemeh.Bayat-B-v2.docx.

Design goals
------------
* fits A4 landscape, 1 page wide, readable when printed
* one row per footnote, RTL sheet, Persian digits avoided in data cells so
  the file stays sortable/filterable
* column «تصمیم» is a dropdown the author fills in:
      اعمال شود    apply the proposed footnote
      بدون تغییر   keep exactly what is there now
      حذف شود      delete this footnote
* the file is re-readable: apply_decisions.py consumes the same columns, so
  the author can print it, mark it up, save it and hand it straight back.
"""

from __future__ import annotations

import json
import os
import re
import sys
import zipfile

from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from sources import SOURCES  # noqa: E402

W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
w = lambda t: f"{{{W}}}{t}"  # noqa: E731

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(REPO, "Payannameh_Fatemeh.Bayat-B-v2.docx")
OUT = os.path.join(REPO, "تصمیم_پانویس_ها.xlsx")

FA2EN = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
HEAD_RE = re.compile(r"^\s*[۰-۹0-9]+(?:\s*-\s*[۰-۹0-9]+)*\s*-\s*\S")
CH_RE = re.compile(r"^\s*فصل\s")


def harvest():
    """Pull every footnote together with where it sits in the thesis."""
    z = zipfile.ZipFile(SRC)
    doc = etree.fromstring(z.read("word/document.xml"))
    fnx = etree.fromstring(z.read("word/footnotes.xml"))

    notes = {}
    for n in fnx.findall(w("footnote")):
        if n.get(w("type")):
            continue
        notes[int(n.get(w("id")))] = "".join(
            t.text or "" for t in n.iter(w("t"))).strip()

    rows, chap, head = {}, "", ""
    for p in doc.iter(w("p")):
        stream = []
        for node in p.iter():
            tag = etree.QName(node).localname
            if tag == "t":
                stream.append(("t", node.text or ""))
            elif tag == "footnoteReference":
                stream.append(("f", int(node.get(w("id")))))
        txt = "".join(s for k, s in stream if k == "t")
        st = txt.strip()
        if st and len(st) < 90:
            if CH_RE.match(st) and len(st) < 40:
                chap = st
            elif HEAD_RE.match(st):
                head = st

        pos, marks = 0, []
        for k, v in stream:
            if k == "t":
                pos += len(v)
            else:
                marks.append((v, pos))

        for fid, at in marks:
            if fid not in notes or fid in rows:
                continue
            before, after = txt[max(0, at - 80):at], txt[at:at + 60]
            m = re.match(r"\s*[\(\)]?\s*([۰-۹0-9]{4})", after)
            year = m.group(1).translate(FA2EN) if m else ""
            if not year:
                mb = re.search(r"[،,]\s*([۰-۹0-9]{4})\s*[؛;\)]?\s*$", before)
                year = mb.group(1).translate(FA2EN) if mb else ""
            term = re.split(r"[(),؛\n]", before.strip())[-1].strip()
            term = " ".join(term.split()[-6:])
            rows[fid] = dict(id=fid, chap=chap, head=head, term=term,
                             year=year, cur=notes[fid])
    return [rows[k] for k in sorted(rows)]


# ---------------------------------------------------------------- styling
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name="B Nazanin", size=11, bold=True, color="FFFFFF")
CELL = Font(name="B Nazanin", size=10)
CELL_LAT = Font(name="Times New Roman", size=10)
CELL_B = Font(name="B Nazanin", size=10, bold=True)
FILL_FIX = PatternFill("solid", fgColor="FCE4E4")   # needs correcting
FILL_ASK = PatternFill("solid", fgColor="FFF2CC")   # author must decide
FILL_OK = PatternFill("solid", fgColor="EAF3EA")    # already fine
FILL_OPT = PatternFill("solid", fgColor="E8EEF7")   # optional enrichment

COLS = [
    ("ش",              5,  "شمارهٔ پانویس در فایل ورد"),
    ("فصل",           11,  "فصل"),
    ("عنوان بخش",     26,  "زیرعنوانی که پانویس در آن آمده"),
    ("واژه در متن",   26,  "عبارت فارسی درست پیش از شمارهٔ پانویس"),
    ("سال",            7,  "سالی که در متن ذکر شده"),
    ("پانویس فعلی",   30,  "آنچه اکنون در پایان‌نامه چاپ می‌شود"),
    ("پانویس پیشنهادی", 34, "پیشنهاد ما برای جایگزینی"),
    ("وضعیت",         14,  "نتیجهٔ بررسی"),
    ("تصمیم شما",     14,  "این ستون را پر کنید"),
    ("توضیح شما",     22,  "در صورت نیاز یادداشت بنویسید"),
]

STATUS_FA = {
    "fixed":      "نیازمند اصلاح",
    "confirmed":  "درست است",
    "unresolved": "منبع پیدا نشد",
}


def build():
    rows = harvest()
    wb = Workbook()

    # ------------------------------------------------ sheet 1: decisions
    ws = wb.active
    ws.title = "تصمیم پانویس‌ها"
    ws.sheet_view.rightToLeft = True

    ws.append([c[0] for c in COLS])
    for i, (_, width, note) in enumerate(COLS, start=1):
        L = get_column_letter(i)
        ws.column_dimensions[L].width = width
        c = ws.cell(row=1, column=i)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.comment = None
    ws.row_dimensions[1].height = 34

    n_fix = n_ask = 0
    for r in rows:
        s = SOURCES.get(r["id"])
        if s:
            proposed = s["short"] or ""
            status = STATUS_FA[s["status"]]
            uncertain = bool(s.get("uncertain"))
        else:
            proposed = ""            # nothing to change
            status = "درست است"
            uncertain = False

        # a proposal that equals the current text is not a change
        if proposed and proposed.strip() == r["cur"].strip():
            proposed = ""

        # Make the status say exactly what the row asks of the reader.
        #   - wrong text            -> must be corrected
        #   - text fine, year added -> optional enrichment
        #   - nothing proposed      -> nothing to do
        if s and s["status"] == "unresolved":
            status = "منبع پیدا نشد"
        elif s and s["status"] == "fixed":
            status = "نیازمند اصلاح"
        elif proposed:
            status = "افزودن سال (اختیاری)"
        else:
            status = "درست است"

        default = ""
        if status == "نیازمند اصلاح":
            default = "اعمال شود"
            n_fix += 1
        if uncertain or status == "منبع پیدا نشد":
            default = ""
            n_ask += 1

        ws.append([r["id"], r["chap"].replace("فصل ", "").rstrip(":"),
                   r["head"], r["term"], r["year"], r["cur"],
                   proposed, status, default, ""])
        row = ws.max_row
        fill = (FILL_ASK if uncertain or status == "منبع پیدا نشد"
                else FILL_FIX if status == "نیازمند اصلاح"
                else FILL_OPT if status == "افزودن سال (اختیاری)"
                else FILL_OK)
        for col in range(1, len(COLS) + 1):
            c = ws.cell(row=row, column=col)
            c.border = BORDER
            c.fill = fill
            latin = col in (6, 7)
            c.font = (CELL_LAT if latin else CELL)
            if col == 7 and proposed:
                c.font = Font(name="Times New Roman", size=10, bold=True)
            c.alignment = Alignment(
                horizontal="left" if latin else
                ("center" if col in (1, 2, 5, 8, 9) else "right"),
                vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30

    dv = DataValidation(
        type="list",
        formula1='"اعمال شود,بدون تغییر,حذف شود"',
        allow_blank=True, showDropDown=False)
    dv.error = "یکی از سه گزینه را انتخاب کنید."
    dv.prompt = "اعمال شود / بدون تغییر / حذف شود"
    ws.add_data_validation(dv)
    dv.add(f"I2:I{ws.max_row}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{ws.max_row}"

    # print setup — one page wide, landscape, repeat header row
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    # ------------------------------------------- sheet 2: full references
    ws2 = wb.create_sheet("منابع کامل")
    ws2.sheet_view.rightToLeft = True
    ws2.append(["ش پانویس", "پانویس کوتاه",
                "مدخل کامل برای بخش منابع (APA)", "در فهرست منابع هست؟"])
    for i, wdt in enumerate((10, 32, 96, 18), start=1):
        L = get_column_letter(i)
        ws2.column_dimensions[L].width = wdt
        c = ws2.cell(row=1, column=i)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    ws2.row_dimensions[1].height = 32

    for fid, s in sorted(SOURCES.items()):
        if not s.get("full"):
            continue
        ws2.append([fid, s["short"], s["full"],
                    "بله" if s.get("inrefs") else "خیر - باید افزوده شود"])
        row = ws2.max_row
        for col in range(1, 5):
            c = ws2.cell(row=row, column=col)
            c.border = BORDER
            c.font = CELL_LAT if col in (2, 3) else CELL
            c.alignment = Alignment(
                horizontal="left" if col in (2, 3) else "center",
                vertical="center", wrap_text=True)
            if col == 4 and not s.get("inrefs"):
                c.fill = FILL_FIX
        ws2.row_dimensions[row].height = 30
    ws2.freeze_panes = "A2"
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.print_title_rows = "1:1"

    # ------------------------------------------------- sheet 3: guidance
    ws3 = wb.create_sheet("راهنما")
    ws3.sheet_view.rightToLeft = True
    ws3.column_dimensions["A"].width = 118
    guide = [
        ("راهنمای استفاده", True),
        ("", False),
        ("۱) این فایل را چاپ کنید (A4 افقی، یک صفحه عرض تنظیم شده است) و "
         "ستون «تصمیم شما» را پر کنید.", False),
        ("۲) در ستون «تصمیم شما» فقط یکی از این سه گزینه را بنویسید: "
         "«اعمال شود» یا «بدون تغییر» یا «حذف شود». در نسخهٔ دیجیتال، این "
         "ستون کرکره‌ای است.", False),
        ("۳) اگر ستون «پانویس پیشنهادی» خالی است یعنی پانویس فعلی درست است "
         "و نیازی به تغییر ندارد.", False),
        ("۴) پس از تکمیل، همین فایل را ذخیره و دوباره بارگذاری کنید تا "
         "نسخهٔ نهایی پایان‌نامه بر اساس تصمیم‌های شما ساخته شود.", False),
        ("", False),
        ("معنی رنگ‌ها", True),
        ("صورتی: پانویس اشتباه است و پیشنهاد اصلاح دارد.", False),
        ("آبی روشن: پانویس درست است؛ فقط افزودن سال پیشنهاد شده که کاملاً اختیاری است.", False),
        ("زرد: منبع قطعی پیدا نشد؛ تصمیم با شماست (یا اصل منبع را ارائه "
         "کنید یا ارجاع حذف شود).", False),
        ("سبز: پانویس درست است؛ نیازی به کار نیست.", False),
        ("", False),
        ("نکته‌های مهم", True),
        ("سبک پیشنهادی: پانویس کوتاه بماند (نام خانوادگی + سال) و مشخصات "
         "کامل در بخش «منابع» بیاید. برگهٔ «منابع کامل» مدخل‌های آمادهٔ درج "
         "را دارد.", False),
        ("۲۱ منبع در متن ارجاع داده شده‌اند اما در فهرست منابع پایان‌نامه "
         "نیستند؛ در برگهٔ «منابع کامل» با «خیر - باید افزوده شود» مشخص "
         "شده‌اند.", False),
        ("سه خطای سال در متن پیدا شد که تصمیم دربارهٔ آن‌ها با شماست: "
         "تورسون و پاوول ۱۹۸۸ است نه ۱۹۹۸؛ استاوروا ۲۰۱۳ است نه ۲۰۱۲؛ "
         "چالن و همکاران ۲۰۱۱ است نه ۲۰۱۰.", False),
        ("دو مورد حل‌نشده: «زنگ ۲۰۱۱» و «وانگ و ژائو ۲۰۲۰» — با این نام و "
         "سال هیچ منبع معتبری یافت نشد.", False),
    ]
    for text, bold in guide:
        ws3.append([text])
        c = ws3.cell(row=ws3.max_row, column=1)
        c.font = Font(name="B Nazanin", size=12 if bold else 11, bold=bold)
        c.alignment = Alignment(horizontal="right", vertical="top",
                                wrap_text=True)
        ws3.row_dimensions[ws3.max_row].height = 30 if text else 10

    wb.save(OUT)
    print("written:", OUT)
    print(f"rows: {len(rows)} | proposed corrections: {n_fix} | "
          f"need your decision: {n_ask}")


if __name__ == "__main__":
    build()
